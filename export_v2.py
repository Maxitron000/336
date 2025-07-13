"""
Система экспорта v2.0 с цветовой индикацией для военного бота
"""

import os
import pandas as pd
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from database import (
    get_all_users, get_location_logs, get_active_users_by_location,
    get_users_without_location, get_admin_logs
)
from utils import get_current_time, format_datetime, get_locations_list

logger = logging.getLogger(__name__)

# Цвета для Excel
GREEN_FILL = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Светло-зеленый
RED_FILL = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')    # Светло-красный
BLUE_FILL = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')   # Светло-голубой
GRAY_FILL = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')   # Светло-серый

# Стили для заголовков
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

def export_military_summary_excel(file_path: str = None) -> str:
    """Экспорт военной сводки в Excel с цветовой индикацией"""
    if not file_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = f"exports/military_summary_{timestamp}.xlsx"
    
    # Создаем директорию если её нет
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    # Создаем рабочую книгу
    wb = Workbook()
    
    # Удаляем дефолтный лист
    wb.remove(wb.active)
    
    # Лист 1: Общая сводка
    create_summary_sheet(wb, "Общая сводка")
    
    # Лист 2: Детализация по локациям
    create_locations_sheet(wb, "Локации")
    
    # Лист 3: Журнал событий
    create_events_sheet(wb, "Журнал событий")
    
    # Лист 4: Админские действия
    create_admin_sheet(wb, "Админские действия")
    
    # Сохраняем файл
    wb.save(file_path)
    logger.info(f"Экспорт военной сводки сохранен: {file_path}")
    
    return file_path

def create_summary_sheet(wb: Workbook, sheet_name: str):
    """Создание листа общей сводки"""
    ws = wb.create_sheet(sheet_name)
    
    # Заголовки
    headers = [
        "ФИО", "Роль", "Статус", "Текущая локация", 
        "Время прибытия", "Последняя активность", "ID Telegram"
    ]
    
    # Добавляем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Получаем данные
    all_users = get_all_users()
    active_locations = get_active_users_by_location()
    users_without_location = get_users_without_location()
    
    # Создаем словарь текущих локаций
    current_locations = {}
    for location, users in active_locations.items():
        for user in users:
            current_locations[user['telegram_id']] = {
                'location': location,
                'entered_at': user['entered_at']
            }
    
    # Заполняем данные
    row = 2
    for user in all_users:
        user_id = user['telegram_id']
        current_location_data = current_locations.get(user_id)
        
        # Определяем статус
        if current_location_data:
            status = "🟢 В расположении"
            location = current_location_data['location']
            arrival_time = format_datetime(current_location_data['entered_at'], 'short')
            status_fill = GREEN_FILL
        else:
            status = "🔴 Вне расположения"
            location = "Не указано"
            arrival_time = "Не указано"
            status_fill = RED_FILL
        
        # Заполняем строку
        ws.cell(row=row, column=1, value=user['full_name'])
        ws.cell(row=row, column=2, value=get_role_display_name(user.get('role', 'soldier')))
        
        status_cell = ws.cell(row=row, column=3, value=status)
        status_cell.fill = status_fill
        
        ws.cell(row=row, column=4, value=location)
        ws.cell(row=row, column=5, value=arrival_time)
        ws.cell(row=row, column=6, value=format_datetime(user['last_activity'], 'short'))
        ws.cell(row=row, column=7, value=user_id)
        
        row += 1
    
    # Автоподбор ширины столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_locations_sheet(wb: Workbook, sheet_name: str):
    """Создание листа детализации по локациям"""
    ws = wb.create_sheet(sheet_name)
    
    # Заголовки
    headers = ["Локация", "Количество", "ФИО", "Время прибытия", "Роль"]
    
    # Добавляем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Получаем данные
    active_locations = get_active_users_by_location()
    
    # Заполняем данные
    row = 2
    for location, users in active_locations.items():
        for i, user in enumerate(users):
            # Локация (только для первой строки группы)
            if i == 0:
                location_cell = ws.cell(row=row, column=1, value=location)
                location_cell.fill = BLUE_FILL
                
                # Количество (только для первой строки группы)
                count_cell = ws.cell(row=row, column=2, value=len(users))
                count_cell.fill = BLUE_FILL
            else:
                ws.cell(row=row, column=1, value="")
                ws.cell(row=row, column=2, value="")
            
            # Данные пользователя
            ws.cell(row=row, column=3, value=user['full_name'])
            ws.cell(row=row, column=4, value=format_datetime(user['entered_at'], 'short'))
            ws.cell(row=row, column=5, value=get_role_display_name(user.get('role', 'soldier')))
            
            row += 1
    
    # Автоподбор ширины столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_events_sheet(wb: Workbook, sheet_name: str):
    """Создание листа журнала событий"""
    ws = wb.create_sheet(sheet_name)
    
    # Заголовки
    headers = ["Дата/Время", "ФИО", "Действие", "Локация", "ID Telegram"]
    
    # Добавляем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Получаем данные (последние 1000 записей)
    logs = get_location_logs(limit=1000)
    
    # Заполняем данные
    row = 2
    for log in logs:
        # Определяем цвет в зависимости от действия
        if log['action'] == 'arrived':
            action_fill = GREEN_FILL
            action_text = "🟢 Прибыл"
        else:
            action_fill = RED_FILL
            action_text = "🔴 Убыл"
        
        # Заполняем строку
        ws.cell(row=row, column=1, value=format_datetime(log['timestamp'], 'full'))
        ws.cell(row=row, column=2, value=log['full_name'])
        
        action_cell = ws.cell(row=row, column=3, value=action_text)
        action_cell.fill = action_fill
        
        ws.cell(row=row, column=4, value=log['location'])
        ws.cell(row=row, column=5, value=log['telegram_id'])
        
        row += 1
    
    # Автоподбор ширины столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_admin_sheet(wb: Workbook, sheet_name: str):
    """Создание листа админских действий"""
    ws = wb.create_sheet(sheet_name)
    
    # Заголовки
    headers = ["Дата/Время", "Админ", "Действие", "Цель", "Детали"]
    
    # Добавляем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Получаем данные (последние 500 записей)
    admin_logs = get_admin_logs(limit=500)
    
    # Заполняем данные
    row = 2
    for log in admin_logs:
        # Заполняем строку
        ws.cell(row=row, column=1, value=format_datetime(log['timestamp'], 'full'))
        ws.cell(row=row, column=2, value=log['admin_name'])
        ws.cell(row=row, column=3, value=log['action'])
        ws.cell(row=row, column=4, value=log.get('target_name', ''))
        ws.cell(row=row, column=5, value=log.get('details', ''))
        
        row += 1
    
    # Автоподбор ширины столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def get_role_display_name(role: str) -> str:
    """Получение отображаемого названия роли"""
    role_names = {
        'soldier': 'Боец',
        'admin': 'Админ',
        'main_admin': 'Главный админ'
    }
    return role_names.get(role, 'Неизвестно')

# Функции для обратной совместимости
def export_data(export_type: str, **kwargs) -> str:
    """Экспорт данных (обратная совместимость)"""
    if export_type == 'excel':
        return export_military_summary_excel()
    else:
        # Для других типов используем старую систему
        from export import export_data as old_export_data
        return old_export_data(export_type, **kwargs)