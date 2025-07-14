"""
Система экспорта данных для бота учета персонала
"""

import os
import csv
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Optional
import pandas as pd
from io import BytesIO

from database import get_location_logs, get_all_users, get_admin_logs, get_active_users_by_location
from utils import get_current_time, format_datetime, get_export_path

logger = logging.getLogger(__name__)

async def export_data(format_type: str, period: Optional[str] = None, user_filter: Optional[str] = None) -> Optional[str]:
    """
    Экспорт данных в указанном формате
    
    Args:
        format_type: Формат экспорта (csv, excel, pdf)
        period: Период для экспорта (day, week, month, all)
        user_filter: Фильтр по пользователю
    
    Returns:
        Путь к созданному файлу или None в случае ошибки
    """
    try:
        # Получаем данные для экспорта
        logs_data = get_location_logs(limit=10000, user_filter=user_filter)
        users_data = get_all_users()
        
        if not logs_data:
            logger.warning("Нет данных для экспорта")
            return None
        
        # Создаем директорию для экспорта
        export_dir = get_export_path()
        os.makedirs(export_dir, exist_ok=True)
        
        # Генерируем имя файла
        timestamp = get_current_time().strftime("%Y%m%d_%H%M%S")
        filename = f"personnel_export_{timestamp}"
        
        if format_type == "csv":
            return await export_to_csv(logs_data, users_data, export_dir, filename)
        elif format_type == "excel":
            return await export_to_excel(logs_data, users_data, export_dir, filename)
        elif format_type == "pdf":
            return await export_to_pdf(logs_data, users_data, export_dir, filename)
        else:
            logger.error(f"Неподдерживаемый формат экспорта: {format_type}")
            return None
            
    except Exception as e:
        logger.error(f"Ошибка при экспорте данных: {e}")
        return None

async def export_to_csv(logs_data: List[Dict], users_data: List[Dict], export_dir: str, filename: str) -> str:
    """Экспорт данных в CSV формат"""
    try:
        filepath = os.path.join(export_dir, f"{filename}.csv")
        
        # Подготавливаем данные для экспорта
        export_data = []
        
        for log in logs_data:
            export_data.append({
                'Дата и время': format_datetime(log['timestamp']),
                'ФИО': log['full_name'],
                'Telegram ID': log['telegram_id'],
                'Действие': 'Прибыл' if log['action'] == 'arrived' else 'Покинул',
                'Локация': log['location']
            })
        
        # Записываем в CSV
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['Дата и время', 'ФИО', 'Telegram ID', 'Действие', 'Локация']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            writer.writerows(export_data)
        
        logger.info(f"Экспорт CSV завершен: {filepath}")
        return filepath
        
    except Exception as e:
        logger.error(f"Ошибка при экспорте в CSV: {e}")
        raise

async def export_to_excel(logs_data: List[Dict], users_data: List[Dict], export_dir: str, filename: str) -> str:
    """Экспорт данных в Excel формат с цветовой заливкой"""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        filepath = os.path.join(export_dir, f"{filename}.xlsx")
        
        # Создаем Excel файл
        from openpyxl import Workbook
        wb = Workbook()
        
        # Определяем стили
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Светло-зеленый
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")   # Светло-красный
        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Светло-голубой
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Синий заголовок
        
        header_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), 
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        # === Лист 1: Журнал действий ===
        ws_journal = wb.active
        ws_journal.title = "Журнал действий"
        
        # Заголовки
        headers = ['Дата и время', 'ФИО', 'Telegram ID', 'Действие', 'Локация']
        for col, header in enumerate(headers, 1):
            cell = ws_journal.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Данные журнала
        for row_idx, log in enumerate(logs_data, 2):
            action_text = 'Прибыл' if log['action'] == 'arrived' else 'Убыл'
            
            # Заполняем данные
            ws_journal.cell(row=row_idx, column=1, value=format_datetime(log['timestamp']))
            ws_journal.cell(row=row_idx, column=2, value=log['full_name'])
            ws_journal.cell(row=row_idx, column=3, value=log['telegram_id'])
            ws_journal.cell(row=row_idx, column=4, value=action_text)
            ws_journal.cell(row=row_idx, column=5, value=log['location'])
            
            # Применяем цветовую заливку в зависимости от действия
            fill_color = green_fill if log['action'] == 'arrived' else red_fill
            
            for col in range(1, 6):
                cell = ws_journal.cell(row=row_idx, column=col)
                cell.fill = fill_color
                cell.border = thin_border
                cell.alignment = center_alignment
        
        # Автоподбор ширины колонок
        for col in range(1, 6):
            ws_journal.column_dimensions[chr(64 + col)].width = 20
        
        # === Лист 2: Пользователи ===
        ws_users = wb.create_sheet("Пользователи")
        
        # Заголовки
        user_headers = ['ФИО', 'Telegram ID', 'Username', 'Администратор', 'Дата регистрации', 'Последняя активность']
        for col, header in enumerate(user_headers, 1):
            cell = ws_users.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Данные пользователей
        for row_idx, user in enumerate(users_data, 2):
            ws_users.cell(row=row_idx, column=1, value=user.get('full_name', user.get('name', '')))
            ws_users.cell(row=row_idx, column=2, value=user['telegram_id'])
            ws_users.cell(row=row_idx, column=3, value=user.get('username', ''))
            ws_users.cell(row=row_idx, column=4, value='Да' if user.get('is_admin') else 'Нет')
            ws_users.cell(row=row_idx, column=5, value=format_datetime(user.get('created_at', '')) if user.get('created_at') else '')
            ws_users.cell(row=row_idx, column=6, value=format_datetime(user.get('updated_at', '')) if user.get('updated_at') else '')
            
            # Выделяем админов голубым цветом
            fill_color = blue_fill if user.get('is_admin') else None
            
            for col in range(1, 7):
                cell = ws_users.cell(row=row_idx, column=col)
                if fill_color:
                    cell.fill = fill_color
                cell.border = thin_border
                cell.alignment = center_alignment
        
        # Автоподбор ширины колонок
        for col in range(1, 7):
            ws_users.column_dimensions[chr(64 + col)].width = 18
        
        # === Лист 3: Статистика ===
        ws_stats = wb.create_sheet("Статистика")
        
        # Общая статистика
        stats = [
            ["Показатель", "Значение"],
            ["Всего записей в журнале", len(logs_data)],
            ["Всего пользователей", len(users_data)],
            ["Администраторов", sum(1 for user in users_data if user.get('is_admin'))],
            ["Обычных пользователей", len(users_data) - sum(1 for user in users_data if user.get('is_admin'))],
            ["", ""],  # Пустая строка
            ["Действия", "Количество"],
        ]
        
        # Считаем действия
        actions_count = {}
        for log in logs_data:
            action = 'Прибытия' if log['action'] == 'arrived' else 'Убытия'
            actions_count[action] = actions_count.get(action, 0) + 1
        
        for action, count in actions_count.items():
            stats.append([action, count])
        
        # Записываем статистику
        for row_idx, (key, value) in enumerate(stats, 1):
            ws_stats.cell(row=row_idx, column=1, value=key)
            ws_stats.cell(row=row_idx, column=2, value=value)
            
            # Выделяем заголовки
            if row_idx == 1 or row_idx == 7:
                for col in range(1, 3):
                    cell = ws_stats.cell(row=row_idx, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.border = thin_border
            else:
                for col in range(1, 3):
                    cell = ws_stats.cell(row=row_idx, column=col)
                    cell.border = thin_border
                    cell.alignment = center_alignment
        
        # Автоподбор ширины колонок
        ws_stats.column_dimensions['A'].width = 25
        ws_stats.column_dimensions['B'].width = 15
        
        # Сохраняем файл
        wb.save(filepath)
        
        logger.info(f"Экспорт Excel с цветовой заливкой завершен: {filepath}")
        return filepath
        
    except Exception as e:
        logger.error(f"Ошибка при экспорте в Excel: {e}")
        raise

async def export_to_pdf(logs_data: List[Dict], users_data: List[Dict], export_dir: str, filename: str) -> str:
    """Экспорт данных в PDF формат"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        
        filepath = os.path.join(export_dir, f"{filename}.pdf")
        
        # Создаем PDF документ
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Заголовок
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
        )
        
        title = Paragraph("Отчет по учету персонала", title_style)
        story.append(title)
        
        # Информация о создании отчета
        info_style = ParagraphStyle(
            'Info',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=20,
        )
        
        current_time = get_current_time()
        info_text = f"Отчет создан: {current_time.strftime('%d.%m.%Y %H:%M')}<br/>Всего записей: {len(logs_data)}<br/>Пользователей в системе: {len(users_data)}"
        info = Paragraph(info_text, info_style)
        story.append(info)
        
        # Таблица с журналом действий
        story.append(Paragraph("Журнал действий", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Подготавливаем данные для таблицы
        table_data = [['Дата и время', 'ФИО', 'Действие', 'Локация']]
        
        for log in logs_data[:50]:  # Ограничиваем количество записей для PDF
            table_data.append([
                format_datetime(log['timestamp'], 'short'),
                log['full_name'],
                'Прибыл' if log['action'] == 'arrived' else 'Покинул',
                log['location']
            ])
        
        # Создаем таблицу
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        
        # Статистика
        story.append(Spacer(1, 20))
        story.append(Paragraph("Статистика", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Получаем активные локации
        active_locations = get_active_users_by_location()
        
        if active_locations:
            stats_text = "<b>Текущее распределение по локациям:</b><br/>"
            for location, users in active_locations.items():
                stats_text += f"• {location}: {len(users)} чел.<br/>"
        else:
            stats_text = "Все локации пусты"
        
        stats = Paragraph(stats_text, styles['Normal'])
        story.append(stats)
        
        # Создаем PDF
        doc.build(story)
        
        logger.info(f"Экспорт PDF завершен: {filepath}")
        return filepath
        
    except Exception as e:
        logger.error(f"Ошибка при экспорте в PDF: {e}")
        raise

def generate_statistics_data(logs_data: List[Dict], users_data: List[Dict]) -> List[Dict]:
    """Генерация данных для статистики"""
    try:
        # Подсчитываем статистику по действиям
        actions_count = {}
        locations_count = {}
        users_activity = {}
        
        for log in logs_data:
            # Статистика по действиям
            action = 'Прибытие' if log['action'] == 'arrived' else 'Убытие'
            actions_count[action] = actions_count.get(action, 0) + 1
            
            # Статистика по локациям
            location = log['location']
            locations_count[location] = locations_count.get(location, 0) + 1
            
            # Статистика по пользователям
            user = log['full_name']
            users_activity[user] = users_activity.get(user, 0) + 1
        
        # Формируем данные для таблицы
        stats_data = []
        
        # Добавляем статистику по действиям
        stats_data.append({'Категория': 'Действия', 'Название': '', 'Количество': ''})
        for action, count in actions_count.items():
            stats_data.append({'Категория': '', 'Название': action, 'Количество': count})
        
        # Добавляем статистику по локациям
        stats_data.append({'Категория': 'Локации', 'Название': '', 'Количество': ''})
        for location, count in sorted(locations_count.items(), key=lambda x: x[1], reverse=True):
            stats_data.append({'Категория': '', 'Название': location, 'Количество': count})
        
        # Добавляем топ активных пользователей
        stats_data.append({'Категория': 'Активные пользователи', 'Название': '', 'Количество': ''})
        for user, count in sorted(users_activity.items(), key=lambda x: x[1], reverse=True)[:10]:
            stats_data.append({'Категория': '', 'Название': user, 'Количество': count})
        
        return stats_data
        
    except Exception as e:
        logger.error(f"Ошибка при генерации статистики: {e}")
        return []

async def export_admin_logs(format_type: str) -> Optional[str]:
    """Экспорт административных логов"""
    try:
        admin_logs = get_admin_logs(limit=1000)
        
        if not admin_logs:
            logger.warning("Нет административных логов для экспорта")
            return None
        
        export_dir = get_export_path()
        os.makedirs(export_dir, exist_ok=True)
        
        timestamp = get_current_time().strftime("%Y%m%d_%H%M%S")
        filename = f"admin_logs_{timestamp}"
        
        if format_type == "csv":
            filepath = os.path.join(export_dir, f"{filename}.csv")
            
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['Дата и время', 'Администратор', 'Действие', 'Цель', 'Детали']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                
                writer.writeheader()
                for log in admin_logs:
                    writer.writerow({
                        'Дата и время': format_datetime(log['timestamp']),
                        'Администратор': log['admin_name'],
                        'Действие': log['action'],
                        'Цель': log.get('target_name', ''),
                        'Детали': log.get('details', '')
                    })
            
            return filepath
            
        elif format_type == "excel":
            filepath = os.path.join(export_dir, f"{filename}.xlsx")
            
            df = pd.DataFrame([
                {
                    'Дата и время': format_datetime(log['timestamp']),
                    'Администратор': log['admin_name'],
                    'Действие': log['action'],
                    'Цель': log.get('target_name', ''),
                    'Детали': log.get('details', '')
                }
                for log in admin_logs
            ])
            
            df.to_excel(filepath, index=False)
            return filepath
        
        return None
        
    except Exception as e:
        logger.error(f"Ошибка при экспорте административных логов: {e}")
        return None

async def export_users_data(format_type: str) -> Optional[str]:
    """Экспорт данных пользователей"""
    try:
        users_data = get_all_users()
        active_locations = get_active_users_by_location()
        
        if not users_data:
            logger.warning("Нет пользователей для экспорта")
            return None
        
        export_dir = get_export_path()
        os.makedirs(export_dir, exist_ok=True)
        
        timestamp = get_current_time().strftime("%Y%m%d_%H%M%S")
        filename = f"users_data_{timestamp}"
        
        # Подготавливаем данные с текущими локациями
        users_with_locations = []
        for user in users_data:
            current_location = None
            for location, location_users in active_locations.items():
                if any(u['telegram_id'] == user['telegram_id'] for u in location_users):
                    current_location = location
                    break
            
            users_with_locations.append({
                'ФИО': user['full_name'],
                'Telegram ID': user['telegram_id'],
                'Username': user.get('username', ''),
                'Администратор': 'Да' if user.get('is_admin') else 'Нет',
                'Текущая локация': current_location or 'Не указана',
                'Дата регистрации': format_datetime(user['registered_at']) if user.get('registered_at') else '',
                'Последняя активность': format_datetime(user['last_activity']) if user.get('last_activity') else ''
            })
        
        if format_type == "csv":
            filepath = os.path.join(export_dir, f"{filename}.csv")
            
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['ФИО', 'Telegram ID', 'Username', 'Администратор', 'Текущая локация', 'Дата регистрации', 'Последняя активность']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                
                writer.writeheader()
                writer.writerows(users_with_locations)
            
            return filepath
            
        elif format_type == "excel":
            filepath = os.path.join(export_dir, f"{filename}.xlsx")
            
            df = pd.DataFrame(users_with_locations)
            df.to_excel(filepath, index=False)
            return filepath
        
        return None
        
    except Exception as e:
        logger.error(f"Ошибка при экспорте данных пользователей: {e}")
        return None

async def cleanup_old_exports(days_to_keep: int = 7):
    """Очистка старых файлов экспорта"""
    try:
        export_dir = get_export_path()
        
        if not os.path.exists(export_dir):
            return
        
        current_time = get_current_time()
        cutoff_time = current_time - timedelta(days=days_to_keep)
        
        deleted_count = 0
        for filename in os.listdir(export_dir):
            filepath = os.path.join(export_dir, filename)
            
            if os.path.isfile(filepath):
                file_time = datetime.fromtimestamp(os.path.getmtime(filepath))
                
                if file_time < cutoff_time:
                    os.remove(filepath)
                    deleted_count += 1
        
        logger.info(f"Очищено {deleted_count} старых файлов экспорта")
        
    except Exception as e:
        logger.error(f"Ошибка при очистке старых экспортов: {e}")

def get_export_info() -> Dict:
    """Получение информации о файлах экспорта"""
    try:
        export_dir = get_export_path()
        
        if not os.path.exists(export_dir):
            return {'files': [], 'total_size': 0}
        
        files = []
        total_size = 0
        
        for filename in os.listdir(export_dir):
            filepath = os.path.join(export_dir, filename)
            
            if os.path.isfile(filepath):
                file_size = os.path.getsize(filepath)
                file_time = datetime.fromtimestamp(os.path.getmtime(filepath))
                
                files.append({
                    'name': filename,
                    'size': file_size,
                    'created': file_time.strftime('%d.%m.%Y %H:%M'),
                    'path': filepath
                })
                
                total_size += file_size
        
        # Сортируем по времени создания (новые первыми)
        files.sort(key=lambda x: x['created'], reverse=True)
        
        return {
            'files': files,
            'total_size': total_size,
            'count': len(files)
        }
        
    except Exception as e:
        logger.error(f"Ошибка при получении информации об экспортах: {e}")
        return {'files': [], 'total_size': 0, 'count': 0}